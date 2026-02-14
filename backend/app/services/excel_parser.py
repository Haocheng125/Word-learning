import openpyxl
import logging
import re

logger = logging.getLogger(__name__)

def parse_excel(excel_path):
    word_dict = {}
    try:
        workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        sheet = workbook.active
        
        logger.info(f'解析Excel: {sheet.title}, {sheet.max_row}行, {sheet.max_column}列')
        
        for row_idx in range(1, sheet.max_row + 1):
            try:
                row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
                
                cells = []
                for cell in row:
                    if cell is None:
                        cells.append('')
                    else:
                        cells.append(str(cell).strip())
                
                # 处理 Excel 读取
                if len(cells) >= 3:
                    # 3 列：序号、英文、中文
                    serial = cells[0]
                    word_str = cells[1]
                    meaning_str = cells[2]
                    
                    if serial and serial.isdigit() and word_str and meaning_str:
                        serial_num = int(serial)
                        
                        word = ''
                        phonetic = ''
                        phonetic_match = re.search(r'\[([^\]]+)\]', word_str)
                        if phonetic_match:
                            word = word_str.replace(phonetic_match.group(0), '').strip()
                            phonetic = phonetic_match.group(1).strip()
                        else:
                            word = word_str
                        
                        if word and serial_num not in word_dict:
                            word_dict[serial_num] = (word, phonetic, meaning_str)
                elif len(cells) >= 2:
                    # 2 列：英文、中文（没有序号，用行号当序号）
                    word_str = cells[0]
                    meaning_str = cells[1]
                    
                    if word_str and meaning_str:
                        serial_num = row_idx
                        
                        word = ''
                        phonetic = ''
                        phonetic_match = re.search(r'\[([^\]]+)\]', word_str)
                        if phonetic_match:
                            word = word_str.replace(phonetic_match.group(0), '').strip()
                            phonetic = phonetic_match.group(1).strip()
                        else:
                            word = word_str
                        
                        if word and serial_num not in word_dict:
                            word_dict[serial_num] = (word, phonetic, meaning_str)
            
            except Exception as e:
                logger.warning(f'解析第{row_idx}行失败: {e}')
        
        sorted_words = []
        if word_dict:
            max_serial = max(word_dict.keys())
            for i in range(1, max_serial + 1):
                if i in word_dict:
                    sorted_words.append(word_dict[i])
        
        logger.info(f'完成，共{len(sorted_words)}个单词')
        
    except FileNotFoundError:
        raise Exception(f'找不到文件: {excel_path}')
    except Exception as e:
        logger.error(f'解析失败: {e}', exc_info=True)
        raise Exception(f'解析失败: {e}')
    
    return sorted_words
